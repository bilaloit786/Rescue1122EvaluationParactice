import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from typing import List, Dict, Any


def generate_excel_report(attempts: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"

    header_fill = PatternFill(start_color="185FA5", end_color="185FA5", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    pass_fill = PatternFill(start_color="EAF3DE", end_color="EAF3DE", fill_type="solid")
    fail_fill = PatternFill(start_color="FAECE7", end_color="FAECE7", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    headers = [
        "Sr#", "Staff Name", "Father Name", "Designation", "District",
        "Station", "Employee ID", "Topic", "Score (%)", "Correct",
        "Total Qs", "Result", "Time (min)", "Date & Time"
    ]
    col_widths = [5, 20, 20, 18, 15, 15, 14, 25, 12, 10, 10, 10, 12, 20]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    ws.row_dimensions[1].height = 22

    for row_idx, attempt in enumerate(attempts, start=2):
        profile = attempt.get("user", {}).get("profile", {}) or {}
        passed = attempt.get("passed", False)
        row_fill = pass_fill if passed else fail_fill
        values = [
            row_idx - 1,
            profile.get("full_name", ""),
            profile.get("father_name", ""),
            profile.get("designation", ""),
            profile.get("district", ""),
            profile.get("station", ""),
            profile.get("employee_id", ""),
            attempt.get("topic_label", ""),
            round(attempt.get("score_percent", 0), 1),
            attempt.get("correct_answers", 0),
            attempt.get("total_questions", 25),
            "PASSED" if passed else "FAILED",
            round((attempt.get("time_taken_seconds") or 0) / 60, 1),
            attempt.get("completed_at", "")[:19].replace("T", " ") if attempt.get("completed_at") else "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if col_idx in [1, 9, 10, 11, 12, 13] else "left", vertical="center")
            if col_idx == 12:
                cell.font = Font(bold=True, color="3B6D11" if passed else "993C1D")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).column_letter}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_pdf_report(attempts: List[Dict[str, Any]], title: str = "Staff Evaluation Report") -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle("title", parent=styles["Heading1"],
                                 textColor=colors.HexColor("#185FA5"), fontSize=16, spaceAfter=6)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"],
                               textColor=colors.gray, fontSize=10, spaceAfter=16)
    elements.append(Paragraph("Rescue 1122 — Punjab Emergency Service", title_style))
    elements.append(Paragraph(f"{title} | Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}", sub_style))

    headers = ["Sr#", "Name", "Designation", "District", "Topic", "Score%", "Result", "Date"]
    data = [headers]
    for i, attempt in enumerate(attempts, 1):
        profile = attempt.get("user", {}).get("profile", {}) or {}
        passed = attempt.get("passed", False)
        data.append([
            str(i),
            profile.get("full_name", "")[:25],
            profile.get("designation", "")[:18],
            profile.get("district", ""),
            attempt.get("topic_label", "")[:28],
            f"{attempt.get('score_percent', 0):.0f}%",
            "PASSED" if passed else "FAILED",
            attempt.get("completed_at", "")[:10] if attempt.get("completed_at") else "",
        ])

    col_widths = [1.2*cm, 5*cm, 4*cm, 3.5*cm, 6.5*cm, 2*cm, 2.5*cm, 2.5*cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#185FA5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F8F8")]),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("ROWHEIGHT", (0, 0), (-1, -1), 18),
    ]))
    elements.append(table)

    doc.build(elements)
    buf.seek(0)
    return buf.read()


def generate_staff_pdf_report(staff: Dict[str, Any], attempts: List[Dict[str, Any]], stats: Dict[str, Any]) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=1.6 * cm,
        rightMargin=1.6 * cm,
        topMargin=1.8 * cm,
        bottomMargin=1.8 * cm,
    )
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle(
        "staff_title",
        parent=styles["Heading1"],
        textColor=colors.HexColor("#185FA5"),
        fontSize=17,
        leading=21,
        spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        "staff_sub",
        parent=styles["Normal"],
        textColor=colors.HexColor("#5F5E5A"),
        fontSize=9,
        spaceAfter=14,
    )
    section_style = ParagraphStyle(
        "section",
        parent=styles["Heading2"],
        textColor=colors.HexColor("#042C53"),
        fontSize=11,
        spaceBefore=12,
        spaceAfter=8,
    )

    profile = staff.get("profile") or {}
    staff_name = profile.get("full_name") or staff.get("username") or "Staff Member"

    elements.append(Paragraph("Rescue 1122 Staff Login Detail", title_style))
    elements.append(Paragraph(
        f"{staff_name} | Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}",
        sub_style,
    ))

    profile_data = [
        ["Full Name", staff_name, "Username", staff.get("username", "")],
        ["Email", staff.get("email", ""), "Status", "Active" if staff.get("is_active") else "Inactive"],
        ["Father Name", profile.get("father_name") or "-", "Designation", profile.get("designation") or "-"],
        ["District", profile.get("district") or "-", "Station", profile.get("station") or "-"],
        ["Employee ID", profile.get("employee_id") or "-", "Phone", profile.get("phone") or "-"],
        ["Joined", (staff.get("created_at") or "")[:10], "Total Tests", str(stats.get("total_tests", 0))],
    ]
    profile_table = Table(profile_data, colWidths=[3.2 * cm, 5.2 * cm, 3.2 * cm, 5.2 * cm])
    profile_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F5F4F0")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E6F1FB")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#E6F1FB")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#185FA5")),
        ("TEXTCOLOR", (2, 0), (2, -1), colors.HexColor("#185FA5")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D3D1C7")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (1, 0), (1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
        ("ROWBACKGROUNDS", (3, 0), (3, -1), [colors.white, colors.HexColor("#FAFAFA")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(profile_table)

    elements.append(Paragraph("Performance Summary", section_style))
    summary_data = [[
        "Average Score",
        "Pass Rate",
        "Passed Tests",
        "Best Score",
    ], [
        f"{stats.get('avg_score', 0):.1f}%",
        f"{stats.get('pass_rate', 0):.1f}%",
        f"{stats.get('passed_tests', 0)} / {stats.get('total_tests', 0)}",
        f"{stats.get('best_score', 0):.1f}%",
    ]]
    summary_table = Table(summary_data, colWidths=[4.4 * cm] * 4)
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#185FA5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#EAF3DE")),
        ("TEXTCOLOR", (0, 1), (-1, 1), colors.HexColor("#1A1A1A")),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D3D1C7")),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)

    elements.append(Paragraph("Recent Test Attempts", section_style))
    attempt_data = [["#", "Topic", "Score", "Correct", "Result", "Date"]]
    for i, attempt in enumerate(attempts[:25], 1):
        passed = attempt.get("passed", False)
        attempt_data.append([
            str(i),
            attempt.get("topic_label", "")[:42],
            f"{attempt.get('score_percent', 0):.0f}%",
            f"{attempt.get('correct_answers', 0)}/{attempt.get('total_questions', 0)}",
            "PASSED" if passed else "FAILED",
            (attempt.get("completed_at") or "")[:16].replace("T", " "),
        ])
    if len(attempt_data) == 1:
        attempt_data.append(["-", "No test attempts recorded", "-", "-", "-", "-"])

    attempts_table = Table(attempt_data, colWidths=[1 * cm, 7 * cm, 2 * cm, 2 * cm, 2.6 * cm, 3.2 * cm], repeatRows=1)
    attempts_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#042C53")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (1, 1), (1, -1), "LEFT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F4F0")]),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D3D1C7")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(attempts_table)

    doc.build(elements)
    buf.seek(0)
    return buf.read()


def generate_attempt_pdf_report(attempt: Dict[str, Any]) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=1.6 * cm,
        rightMargin=1.6 * cm,
        topMargin=1.8 * cm,
        bottomMargin=1.8 * cm,
    )
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle(
        "attempt_title",
        parent=styles["Heading1"],
        textColor=colors.HexColor("#185FA5"),
        fontSize=17,
        leading=21,
        spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        "attempt_sub",
        parent=styles["Normal"],
        textColor=colors.HexColor("#5F5E5A"),
        fontSize=9,
        spaceAfter=14,
    )
    section_style = ParagraphStyle(
        "attempt_section",
        parent=styles["Heading2"],
        textColor=colors.HexColor("#042C53"),
        fontSize=11,
        spaceBefore=12,
        spaceAfter=8,
    )
    body_style = ParagraphStyle(
        "attempt_body",
        parent=styles["Normal"],
        fontSize=8,
        leading=11,
        textColor=colors.HexColor("#444441"),
    )

    profile = attempt.get("user", {}).get("profile", {}) or {}
    staff_name = profile.get("full_name") or attempt.get("user", {}).get("username") or "Staff Member"
    passed = attempt.get("passed", False)

    elements.append(Paragraph("Rescue 1122 Examination Detail Report", title_style))
    elements.append(Paragraph(
        f"{staff_name} | Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}",
        sub_style,
    ))

    summary_data = [
        ["Staff Name", staff_name, "Topic", attempt.get("topic_label", "")],
        ["Designation", profile.get("designation") or "-", "District", profile.get("district") or "-"],
        ["Score", f"{attempt.get('score_percent', 0):.1f}%", "Result", "PASSED" if passed else "FAILED"],
        ["Correct Answers", f"{attempt.get('correct_answers', 0)}/{attempt.get('total_questions', 0)}", "Time Taken", f"{round((attempt.get('time_taken_seconds') or 0) / 60, 1)} min"],
        ["Completed", (attempt.get("completed_at") or "")[:19].replace("T", " "), "Attempt ID", str(attempt.get("id", ""))],
    ]
    summary_table = Table(summary_data, colWidths=[3.3 * cm, 5.4 * cm, 3.3 * cm, 5.4 * cm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E6F1FB")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#E6F1FB")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#185FA5")),
        ("TEXTCOLOR", (2, 0), (2, -1), colors.HexColor("#185FA5")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D3D1C7")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (1, 0), (1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
        ("ROWBACKGROUNDS", (3, 0), (3, -1), [colors.white, colors.HexColor("#FAFAFA")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    elements.append(summary_table)

    subtopic_scores = attempt.get("subtopic_scores") or {}
    if subtopic_scores:
        elements.append(Paragraph("Subtopic Scores", section_style))
        subtopic_data = [["Subtopic", "Correct", "Total", "Percent"]]
        for topic, score in subtopic_scores.items():
            subtopic_data.append([
                str(topic),
                str(score.get("correct", 0)),
                str(score.get("total", 0)),
                f"{score.get('percent', 0):.1f}%",
            ])
        subtopic_table = Table(subtopic_data, colWidths=[9 * cm, 2.5 * cm, 2.5 * cm, 3 * cm], repeatRows=1)
        subtopic_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#042C53")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D3D1C7")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F4F0")]),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(subtopic_table)

    if attempt.get("ai_feedback"):
        elements.append(Paragraph("Evaluation Feedback", section_style))
        feedback = str(attempt["ai_feedback"]).replace("\n", "<br/>")
        elements.append(Paragraph(feedback, body_style))

    doc.build(elements)
    buf.seek(0)
    return buf.read()
